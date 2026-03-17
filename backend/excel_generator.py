"""Excel file generation for financial analysis using openpyxl."""
import os
import uuid
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


HEADER_FONT = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1B4F72", end_color="1B4F72", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="2E86C1", end_color="2E86C1", fill_type="solid")
SUBHEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
DATA_FONT = Font(name="Calibri", size=11)
NUMBER_FORMAT = '#,##0'
PCT_FORMAT = '0.0%'
THIN_BORDER = Border(
    left=Side(style='thin', color='D5D8DC'),
    right=Side(style='thin', color='D5D8DC'),
    top=Side(style='thin', color='D5D8DC'),
    bottom=Side(style='thin', color='D5D8DC')
)
ALT_ROW_FILL = PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid")


def _style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER


def _style_subheader_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = SUBHEADER_FONT
        cell.fill = SUBHEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER


def _style_data_cell(ws, row, col, is_alt=False):
    cell = ws.cell(row=row, column=col)
    cell.font = DATA_FONT
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal='center' if col > 1 else 'left', vertical='center')
    if is_alt:
        cell.fill = ALT_ROW_FILL


def _auto_width(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max(max_length + 4, 14)


def _write_value(ws, row, col, value, is_alt=False, fmt=None):
    cell = ws.cell(row=row, column=col)
    if value is not None:
        cell.value = value
    else:
        cell.value = "N/A"
    _style_data_cell(ws, row, col, is_alt)
    if fmt and value is not None:
        cell.number_format = fmt


def generate_excel(analysis: dict, output_dir: str) -> str:
    """Generate a styled Excel workbook from analysis data.
    
    Returns the filename of the generated file.
    """
    wb = Workbook()
    company = analysis.get("company_name", "Company")
    
    # --- Tab 1: Assumptions ---
    ws_assumptions = wb.active
    ws_assumptions.title = "Assumptions"
    ws_assumptions.sheet_properties.tabColor = "1B4F72"
    
    ws_assumptions.cell(row=1, column=1, value="Assumptions")
    ws_assumptions.cell(row=1, column=2, value="Value")
    _style_header_row(ws_assumptions, 1, 2)
    
    dcf = analysis.get("dcf", {}) or {}
    assumptions = [
        ("Company", company),
        ("Analysis Date", datetime.now().strftime("%Y-%m-%d")),
        ("Currency", "USD (Millions)"),
        ("WACC", f"{dcf.get('wacc', 'N/A')}%"),
        ("Terminal Growth Rate", f"{dcf.get('terminal_growth_rate', 'N/A')}%"),
    ]
    
    for i, (label, value) in enumerate(assumptions, start=2):
        ws_assumptions.cell(row=i, column=1, value=label)
        ws_assumptions.cell(row=i, column=2, value=value)
        _style_data_cell(ws_assumptions, i, 1, i % 2 == 0)
        _style_data_cell(ws_assumptions, i, 2, i % 2 == 0)
    
    _auto_width(ws_assumptions)
    
    # --- Tab 2: Historical Financials ---
    ws_hist = wb.create_sheet("Historical Financials")
    ws_hist.sheet_properties.tabColor = "2E86C1"
    
    extracted = analysis.get("extracted_data", {}) or {}
    metrics = [
        ("Revenue", "revenue"),
        ("Net Income", "net_income"),
        ("EBITDA", "ebitda"),
        ("Gross Margin %", "gross_margin_pct"),
        ("Free Cash Flow", "free_cash_flow"),
        ("Total Debt", "total_debt"),
        ("Cash", "cash"),
        ("Shares Outstanding", "shares_outstanding"),
    ]
    
    years_covered = analysis.get("years_covered", [])
    if not years_covered and extracted:
        for key in extracted:
            items = extracted.get(key, [])
            if items:
                years_covered = sorted(set(item["year"] for item in items if item.get("year")))
                break
    
    ws_hist.cell(row=1, column=1, value="Metric")
    for j, year in enumerate(years_covered):
        ws_hist.cell(row=1, column=j + 2, value=year)
    _style_header_row(ws_hist, 1, len(years_covered) + 1)
    
    for i, (label, key) in enumerate(metrics, start=2):
        ws_hist.cell(row=i, column=1, value=label)
        _style_data_cell(ws_hist, i, 1, i % 2 == 0)
        items = extracted.get(key, []) or []
        year_map = {item["year"]: item["value"] for item in items if item}
        for j, year in enumerate(years_covered):
            val = year_map.get(year)
            fmt = PCT_FORMAT if "pct" in key or "margin" in key.lower() else NUMBER_FORMAT
            if "pct" in key and val is not None:
                val = val / 100
            _write_value(ws_hist, i, j + 2, val, i % 2 == 0, fmt)
    
    _auto_width(ws_hist)
    
    # --- Tab 3: Projections ---
    ws_proj = wb.create_sheet("Projections")
    ws_proj.sheet_properties.tabColor = "27AE60"
    
    projections = analysis.get("projections", {}) or {}
    proj_years = projections.get("years", [])
    
    ws_proj.cell(row=1, column=1, value="Metric")
    for j, year in enumerate(proj_years):
        ws_proj.cell(row=1, column=j + 2, value=year)
    _style_header_row(ws_proj, 1, len(proj_years) + 1)
    
    proj_metrics = [
        ("Revenue", "revenue"),
        ("EBITDA", "ebitda"),
        ("Free Cash Flow", "fcf"),
    ]
    
    for i, (label, key) in enumerate(proj_metrics, start=2):
        ws_proj.cell(row=i, column=1, value=label)
        _style_data_cell(ws_proj, i, 1, i % 2 == 0)
        values = projections.get(key, []) or []
        for j, val in enumerate(values):
            _write_value(ws_proj, i, j + 2, val, i % 2 == 0, NUMBER_FORMAT)
    
    _auto_width(ws_proj)
    
    # --- Tab 4: DCF Valuation ---
    ws_dcf = wb.create_sheet("DCF Valuation")
    ws_dcf.sheet_properties.tabColor = "E74C3C"
    
    ws_dcf.cell(row=1, column=1, value="DCF Valuation")
    ws_dcf.cell(row=1, column=2, value="Value")
    _style_header_row(ws_dcf, 1, 2)
    
    dcf_items = [
        ("WACC", f"{dcf.get('wacc', 'N/A')}%"),
        ("Terminal Growth Rate", f"{dcf.get('terminal_growth_rate', 'N/A')}%"),
        ("Enterprise Value ($M)", dcf.get("enterprise_value")),
        ("Equity Value ($M)", dcf.get("equity_value")),
        ("Implied Share Price ($)", dcf.get("implied_price")),
        ("Upside / Downside %", f"{dcf.get('upside_downside_pct', 'N/A')}%"),
    ]
    
    for i, (label, value) in enumerate(dcf_items, start=2):
        ws_dcf.cell(row=i, column=1, value=label)
        ws_dcf.cell(row=i, column=2, value=value if value is not None else "N/A")
        _style_data_cell(ws_dcf, i, 1, i % 2 == 0)
        _style_data_cell(ws_dcf, i, 2, i % 2 == 0)
    
    _auto_width(ws_dcf)
    
    # Save
    filename = f"{company.replace(' ', '_')}_analysis_{uuid.uuid4().hex[:8]}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    
    return filename
